"""xlsx template builder for the bulk-import flow.

Two surfaces:

  - `build_fields_template_xlsx()` — empty fields-master template.
    Static (doesn't depend on farm state). Sample row included so the
    user has a concrete reference for date / number formats.

  - `build_yields_template_xlsx(field_names)` — yields template
    pre-populated with one row per existing field on the farm. The
    user fills in season + yield + unit per row and saves as CSV.

Both templates use:
  - Bold header row + frozen panes so the headers stay visible while
    scrolling.
  - Data validation dropdowns where the bulk-import backend enforces
    enum values (crop_type, irrigation, fertigation_capable). Excel
    surfaces these as proper dropdowns; LibreOffice respects them too.
  - An "Instructions" sheet at the front so the user reads the
    workflow before opening Sheet1.

Headers match the alias map in `sapling-web/src/lib/import-csv.ts`
exactly so the user's saved-as-CSV uploads cleanly without renaming.
"""
from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# Styling
# ============================================================

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2F5233")  # Sapling deep green
_INSTRUCTION_FONT = Font(bold=True, size=12, color="2F5233")


def _apply_header_row(ws, headers: list[str]) -> None:
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _set_column_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _add_dropdown(
    ws, *, column_letter: str, options: list[str], first_data_row: int = 2,
    last_data_row: int = 1000,
) -> None:
    """Excel data-validation dropdown. Range covers the data area so
    Excel offers the picker on every row the user might fill in."""
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(f"{column_letter}{first_data_row}:{column_letter}{last_data_row}")
    ws.add_data_validation(dv)


# ============================================================
# Fields template
# ============================================================

_FIELDS_HEADERS = [
    "Block",            # required — block name
    "Crop",             # e.g. macadamia, citrus, wheat
    "Cultivar",         # e.g. Beaumont, Navel
    "Crop Type",        # annual | perennial
    "Area ha",          # required for engine builds
    "Tree Age",         # years (perennial only)
    "Pop per ha",       # trees / plants per ha
    "Yield Target",     # numeric
    "Yield Unit",       # t/ha, t NIS/ha
    "Irrigation",       # drip | pivot | micro | flood | none
    "Fertigation",      # yes | no
    "Soil Type",        # free text
    "Notes",            # free text
]
_FIELDS_WIDTHS = [16, 14, 14, 12, 10, 10, 12, 14, 14, 14, 13, 18, 30]
_FIELDS_INSTRUCTIONS = [
    "Bulk-import template — fields / blocks",
    "",
    "How to use:",
    "  1. Fill in the rows below the header. One row = one block.",
    "  2. Required column: Block (the block name).",
    "  3. Strongly recommended: Area ha — programmes can't be built without it.",
    "  4. Crop Type, Irrigation, and Fertigation use dropdowns — pick the closest match.",
    "  5. Save as CSV (File → Save As → CSV UTF-8). Then upload on the bulk import page.",
    "",
    "Notes:",
    "  • Decimal separator: dot or comma both work.",
    "  • Yield Unit examples: t/ha, t NIS/ha, kg/tree.",
    "  • Existing blocks with the same name are skipped by default; toggle to 'Update' on the import page to overwrite blanks.",
]


def build_fields_template_xlsx() -> bytes:
    """Empty fields-master xlsx with one example row."""
    wb = Workbook()

    # Instructions sheet first so it opens on top.
    info = wb.active
    info.title = "Instructions"
    for i, line in enumerate(_FIELDS_INSTRUCTIONS, start=1):
        cell = info.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = _INSTRUCTION_FONT
    info.column_dimensions["A"].width = 100

    ws = wb.create_sheet(title="Fields")
    _apply_header_row(ws, _FIELDS_HEADERS)
    _set_column_widths(ws, _FIELDS_WIDTHS)

    # Dropdown columns — Crop Type (D), Irrigation (J), Fertigation (K).
    _add_dropdown(ws, column_letter="D", options=["annual", "perennial"])
    _add_dropdown(
        ws, column_letter="J",
        options=["drip", "pivot", "micro", "flood", "none"],
    )
    _add_dropdown(ws, column_letter="K", options=["yes", "no"])

    # Example row so format is visible at a glance.
    example = [
        "Blok 1", "macadamia", "Beaumont", "perennial", 4.5, 9, 312, 4.0,
        "t NIS/ha", "micro", "no", "Sandy loam", "9-year-old stand",
    ]
    for col_idx, value in enumerate(example, start=1):
        ws.cell(row=2, column=col_idx, value=value)
    ws.cell(row=2, column=1).font = Font(italic=True, color="888888")
    note = ws.cell(
        row=4, column=1,
        value="↑ Example row — overwrite or delete it before saving.",
    )
    note.font = Font(italic=True, color="888888")

    return _to_bytes(wb)


# ============================================================
# Yields template
# ============================================================

_YIELDS_HEADERS = [
    "Block",         # required — must match an existing field name on the farm
    "Season",        # required — e.g. 2024/25 or 2025
    "Yield",         # required — numeric
    "Yield Unit",    # required — t/ha, t NIS/ha, kg/tree
    "Harvest Date",  # optional — YYYY-MM-DD
    "Notes",         # free text
]
_YIELDS_WIDTHS = [18, 12, 10, 14, 14, 30]


def _yields_instructions(block_count: int) -> list[str]:
    if block_count == 0:
        block_line = (
            "  • This farm has no blocks yet — Block column is left empty. "
            "Add fields first via the Fields tab, then download this "
            "template again to get them pre-populated."
        )
    else:
        block_line = (
            f"  • The Block column is pre-populated with this farm's "
            f"{block_count} existing block{'s' if block_count != 1 else ''}. "
            "Add a row per (block, season) you want to record."
        )
    return [
        "Bulk-import template — yield records",
        "",
        "How to use:",
        "  1. Fill in Season + Yield + Yield Unit for every row you want to import.",
        "  2. Required columns: Block, Season, Yield, Yield Unit.",
        "  3. Save as CSV (File → Save As → CSV UTF-8). Then upload on the bulk import page.",
        "",
        "Notes:",
        block_line,
        "  • Season examples: 2024/25 (split-year crops like macadamia / citrus), 2025 (single-year).",
        "  • Yield Unit examples: t/ha, t NIS/ha, kg/tree.",
        "  • Harvest Date is optional — use YYYY-MM-DD if you have it.",
        "  • You can duplicate a block name across multiple rows for multi-season history.",
    ]


def build_yields_template_xlsx(field_names: Iterable[str]) -> bytes:
    """Yields xlsx pre-populated with one row per existing field. The
    pre-populated rows have a placeholder Yield Unit ('t/ha') so the
    user only needs to fill Season + Yield to get going."""
    names = sorted({n.strip() for n in field_names if n and n.strip()})

    wb = Workbook()
    info = wb.active
    info.title = "Instructions"
    for i, line in enumerate(_yields_instructions(len(names)), start=1):
        cell = info.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = _INSTRUCTION_FONT
    info.column_dimensions["A"].width = 100

    ws = wb.create_sheet(title="Yields")
    _apply_header_row(ws, _YIELDS_HEADERS)
    _set_column_widths(ws, _YIELDS_WIDTHS)

    if names:
        for row_idx, name in enumerate(names, start=2):
            ws.cell(row=row_idx, column=1, value=name)
            # Sensible default unit so the user only has to fill yield + season.
            ws.cell(row=row_idx, column=4, value="t/ha")
    else:
        # No fields yet — drop in one example row so the format is clear.
        example = ["Blok 1", "2024/25", 3.8, "t NIS/ha", "2025-04-15", ""]
        for col_idx, value in enumerate(example, start=1):
            ws.cell(row=2, column=col_idx, value=value)
        ws.cell(row=2, column=1).font = Font(italic=True, color="888888")
        ws.cell(
            row=4, column=1,
            value="↑ Example row — overwrite or delete it before saving.",
        ).font = Font(italic=True, color="888888")

    return _to_bytes(wb)


# ============================================================
# Internals
# ============================================================

def _to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
