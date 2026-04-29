"""Extract soil analysis values from lab report PDFs/photos using Claude vision."""

import base64
import json
import logging
import time
from typing import Any

import anthropic

from app.config import get_settings
from app.supabase_client import get_supabase_admin

logger = logging.getLogger(__name__)

# Sonnet 4.6 — current canonical ID. Lab extraction is structured
# schema output from a known-shape document, so Sonnet is the right
# tier (Opus is overkill for ~R0.65 saved per upload). Adaptive
# thinking + effort=high handle multi-page / hand-annotated reports.
_LAB_EXTRACTOR_MODEL = "claude-sonnet-4-6"

# Anthropic transient errors to retry on. APIError is the parent;
# RateLimitError + APIConnectionError + InternalServerError are the
# specific subclasses we expect from network blips or load.
_RETRYABLE_ERRORS = (
    anthropic.APIConnectionError,
    anthropic.APITimeoutError,
    anthropic.InternalServerError,
    anthropic.RateLimitError,
)

# Soil parameters we expect from lab reports
SOIL_PARAMS = [
    "pH (KCl)", "pH (H2O)", "Resistance (Ohm)", "Organic C (%)",
    "Total N (%)", "NH4-N (mg/kg)", "NO3-N (mg/kg)",
    "P (Bray-1) (mg/kg)", "P (Bray-2) (mg/kg)", "P (Olsen) (mg/kg)",
    "K (mg/kg)", "Ca (mg/kg)", "Mg (mg/kg)", "Na (mg/kg)",
    "S (mg/kg)", "Mn (mg/kg)", "Fe (mg/kg)", "Cu (mg/kg)",
    "Zn (mg/kg)", "B (mg/kg)", "Mo (mg/kg)",
    "CEC (cmol/kg)", "Clay (%)",
    "Ca Saturation (%)", "Mg Saturation (%)", "K Saturation (%)",
    "Na Saturation (%)", "Acid Saturation (%)",
]

# Leaf / foliar parameters — macros in %, micros in mg/kg.
LEAF_PARAMS = [
    "N (%)", "P (%)", "K (%)", "Ca (%)", "Mg (%)", "S (%)", "Na (%)", "Cl (%)",
    "Mn (mg/kg)", "Fe (mg/kg)", "Cu (mg/kg)", "Zn (mg/kg)", "B (mg/kg)",
    "Mo (mg/kg)", "Al (mg/kg)",
]

# Backwards-compat alias — older callers and templates reference EXPECTED_PARAMS.
EXPECTED_PARAMS = SOIL_PARAMS


def _spreadsheet_to_text(file_bytes: bytes, content_type: str) -> str:
    """Render an xlsx / xls / csv as a tab-separated text block for the
    AI. We deliberately don't try to clean / pre-parse the file — labs
    use wildly different column orders, parameter names, units, and
    sometimes mix soil + leaf in the same workbook. Surfacing the raw
    grid (with sheet names) lets Sonnet do the mapping holistically.

    Hard caps to prevent runaway costs from messy multi-sheet workbooks:
      * Per-cell text truncated at 200 chars (long formula text cells).
      * Per-sheet row cap of 500 — labs almost never exceed 100 samples
        per report; anything past 500 is almost certainly noise.
      * Total output capped at 60,000 characters (~15k tokens). If the
        file exceeds this, we raise — caller surfaces a clear error so
        the user trims the workbook rather than silently burning $2 on
        an oversized prompt.
    """
    from io import BytesIO, StringIO

    if content_type in ("text/csv", "text/plain") or content_type.startswith("text/"):
        # CSVs may use comma or semicolon delimiters in SA labs; let
        # the model figure it out — we just hand it the raw text.
        try:
            text = file_bytes.decode("utf-8", errors="replace")
        except Exception:
            text = file_bytes.decode("latin-1", errors="replace")
        if len(text) > _MAX_TEXT_CHARS:
            raise ValueError(
                f"CSV is too large ({len(text):,} chars). Lab CSVs typically "
                f"fit in ≤ {_MAX_TEXT_CHARS:,} chars. Trim to a single "
                f"analysis sheet and re-upload."
            )
        return text

    # xlsx is OOXML; xls is the legacy Excel 97-2003 binary format.
    # Try openpyxl first (handles .xlsx/.xlsm); fall back to xlrd for
    # .xls. Many SA labs still email old-format reports — we want both
    # to "just work" without forcing the user to convert.
    sheet_iter = _iter_sheet_rows_xlsx(file_bytes)
    if sheet_iter is None:
        sheet_iter = _iter_sheet_rows_xls(file_bytes)
    if sheet_iter is None:
        raise ValueError(
            "Couldn't read this file as a spreadsheet. Save it as .xlsx "
            "in Excel (File → Save As → Excel Workbook) or upload the "
            "lab's PDF instead."
        )

    out = StringIO()
    total_chars = 0
    for sheet_name, rows in sheet_iter:
        header = f"--- Sheet: {sheet_name} ---\n"
        out.write(header)
        total_chars += len(header)
        rows_written = 0
        for row in rows:
            if rows_written >= _MAX_ROWS_PER_SHEET:
                msg = (
                    f"[truncated — sheet '{sheet_name}' exceeded "
                    f"{_MAX_ROWS_PER_SHEET} rows]\n"
                )
                out.write(msg)
                total_chars += len(msg)
                break
            cells = []
            for v in row:
                if v is None:
                    cells.append("")
                    continue
                cell = (
                    v.isoformat() if hasattr(v, "isoformat") else str(v)
                )
                if len(cell) > _MAX_CELL_CHARS:
                    cell = cell[:_MAX_CELL_CHARS] + "…"
                cells.append(cell)
            # Skip fully-blank rows so the AI doesn't waste tokens on
            # padding inside the workbook.
            if not any(c.strip() for c in cells):
                continue
            line = "\t".join(cells) + "\n"
            out.write(line)
            total_chars += len(line)
            rows_written += 1
            if total_chars > _MAX_TEXT_CHARS:
                raise ValueError(
                    f"Spreadsheet is too large for lab extraction "
                    f"({total_chars:,}+ chars). Likely cause: multi-sheet "
                    f"workbook with non-lab data (formulas, summaries, raw "
                    f"data dumps). Save just the analysis sheet as its own "
                    f"file and re-upload, or upload the lab's PDF instead."
                )
        out.write("\n")
        total_chars += 1
    return out.getvalue()


def _iter_sheet_rows_xlsx(file_bytes: bytes):
    """Yield (sheet_name, row_iter) for an .xlsx workbook, or None if
    the file isn't OOXML. Returns a generator-of-generators — caller
    handles size caps."""
    from io import BytesIO
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        return None
    return (
        (sheet_name, wb[sheet_name].iter_rows(values_only=True))
        for sheet_name in wb.sheetnames
    )


def _iter_sheet_rows_xls(file_bytes: bytes):
    """Yield (sheet_name, row_iter) for a legacy .xls workbook, or None
    if xlrd can't read it. xlrd's row API is index-based; we wrap it as
    a generator of tuples to match openpyxl's shape."""
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
    except Exception:
        return None

    def _rows(ws):
        for r in range(ws.nrows):
            yield tuple(ws.cell_value(r, c) for c in range(ws.ncols))

    return (
        (wb.sheet_by_index(i).name, _rows(wb.sheet_by_index(i)))
        for i in range(wb.nsheets)
    )


# Hard caps — guard against runaway costs from messy lab xlsx files.
# 60k chars ≈ 15k tokens at typical English/numeric mix; well under
# Sonnet's context but big enough for a 100-block lab report with all
# 28 SOIL_PARAMS columns.
_MAX_TEXT_CHARS = 60_000
_MAX_ROWS_PER_SHEET = 500
_MAX_CELL_CHARS = 200


def _build_extraction_prompt(
    lab_template: dict | None = None,
    mode: str = "soil",
) -> str:
    """Build the system prompt for Claude to extract lab values.

    `mode` selects soil vs leaf parameter set + crop-specific guidance.
    """
    is_leaf = mode == "leaf"
    domain = "leaf/foliar" if is_leaf else "soil"
    params = LEAF_PARAMS if is_leaf else SOIL_PARAMS
    base = (
        f"You are a {domain} lab report data extractor. Extract ALL analysis values "
        "from this lab report image/PDF.\n\n"
        "IMPORTANT: The report may contain MULTIPLE samples (rows in a table). "
        "Each row is a separate sample/block with its own values.\n\n"
        "Return a JSON object with these keys:\n"
        '- "lab_name": the laboratory name (string)\n'
        '- "report_number": report/certificate number if visible (string or null)\n'
        '- "analysis_date": date of analysis if visible (string or null)\n'
        '- "client": client name if visible (string or null)\n'
        f'- "department": e.g. "{ "Leaf" if is_leaf else "Soil" }", "Water" (string or null)\n'
        '- "samples": array of sample objects, one per row/sample in the report\n\n'
        "Each sample object must have:\n"
        '- "sample_id": lab sample number (string)\n'
        '- "block_name": block/field identifier if visible (string or null)\n'
        '- "crop": fruit/crop type if visible (string or null)\n'
        '- "cultivar": cultivar/variety if visible (string or null)\n'
    )
    if is_leaf:
        base += (
            '- "sample_part": leaf age / part sampled (e.g. "spring flush", '
            '"petiole at bloom") if visible (string or null)\n'
        )
    base += (
        '- "values": object mapping parameter names to numeric values\n\n'
        "If there is only ONE sample, still return it as a single-element array.\n\n"
        "For the values object, use these standard parameter names where possible:\n"
    )
    for p in params:
        base += f"  - {p}\n"

    base += (
        "\nRules:\n"
        "- Extract EVERY row/sample in the table — do not skip any\n"
        "- Values must be numeric (no units in the value, just the number)\n"
        "- If a value is given as a range, use the midpoint\n"
        "- If a value shows '<' (below detection), use 0\n"
        "- Map lab-specific parameter names to the standard names above\n"
        '- Add a top-level "confidence": "high" | "medium" | "low" reflecting\n'
        '  how confident you are in the extracted numbers overall (low = blurry/skewed/illegible)\n'
        '- Per-sample, if more than 3 values are illegible/uncertain, set "low_confidence": true on that sample\n'
    )
    if is_leaf:
        base += (
            "- Macros (N, P, K, Ca, Mg, S, Na, Cl) are typically reported in %\n"
            "- Micros (Mn, Fe, Cu, Zn, B, Mo) are typically in mg/kg or ppm\n"
            "- If lab reports leaf macros in mg/kg or ppm by mistake, convert "
            "to % (divide by 10,000)\n"
        )
    else:
        base += (
            "- For mixed soil+leaf reports, only extract the SOIL section\n"
        )
    base += (
        "- Include norms/reference rows if present (mark them with block_name: 'Norm - Low', 'Norm - High' etc.)\n"
        "- Return ONLY valid JSON, no markdown or explanation\n"
    )

    if lab_template and lab_template.get("field_mappings"):
        base += (
            "\nThis lab report is from a known lab. Use these field mappings "
            "to improve accuracy:\n"
        )
        for lab_field, our_field in lab_template["field_mappings"].items():
            base += f'  - "{lab_field}" → "{our_field}"\n'

    return base


def extract_from_document(
    file_bytes: bytes,
    content_type: str,
    lab_name_hint: str | None = None,
    mode: str = "soil",
    user_id: str | None = None,
) -> dict[str, Any]:
    """Extract soil or leaf analysis values from a lab report PDF or image.

    `mode` selects soil vs leaf parameter set. `user_id` is needed for
    the failure-path ai_usage logger (the success path's logger lives
    in the route and passes user.id). Returns dict with keys:
    lab_name, analysis_date, samples, ai_usage.
    """
    settings = get_settings()
    if not settings.anthropic_api_key:
        raise ValueError("ANTHROPIC_API_KEY not configured")

    client = anthropic.Anthropic(api_key=settings.anthropic_api_key)

    # Check for existing lab template — keyed on (lab_name, department) so
    # soil and leaf templates from the same lab don't collide.
    lab_template = None
    if lab_name_hint:
        lab_template = _get_lab_template(lab_name_hint, department=mode)

    prompt = _build_extraction_prompt(lab_template, mode=mode)

    # Build the message content based on file type. Spreadsheets
    # (xlsx / xls / csv) get converted to a tabular text representation
    # server-side and sent as text input — Claude's vision API doesn't
    # natively read spreadsheet binaries. Lab xlsx exports have wildly
    # variable column orders + parameter names + units across labs, so
    # the AI path handles the mapping the same way it does for PDFs.
    spreadsheet_types = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "text/csv",
        "text/plain",
    )
    is_spreadsheet = (
        content_type in spreadsheet_types
        or content_type.startswith("text/")
    )

    media_type: str | None = None
    source_type: str | None = None
    source_key: str | None = None
    spreadsheet_text: str | None = None
    if content_type == "application/pdf":
        media_type = "application/pdf"
        source_type = "base64"
        source_key = "data"
    elif content_type in ("image/jpeg", "image/png", "image/webp", "image/gif"):
        media_type = content_type
        source_type = "base64"
        source_key = "data"
    elif is_spreadsheet:
        spreadsheet_text = _spreadsheet_to_text(file_bytes, content_type)
    else:
        raise ValueError(f"Unsupported file type: {content_type}")

    b64_data = (
        base64.standard_b64encode(file_bytes).decode("utf-8")
        if not is_spreadsheet
        else ""
    )

    # One retry on transient API errors with a short backoff. Lab uploads
    # are user-blocking; a single retry covers the common load-spike +
    # network-blip case without making the user wait too long.
    message = None
    last_err: Exception | None = None
    for attempt in range(2):
        try:
            if is_spreadsheet:
                content_blocks: list[dict[str, Any]] = [
                    {
                        "type": "text",
                        "text": (
                            "The following is a lab report exported as a "
                            "spreadsheet (xlsx / csv). Treat each sheet's "
                            "rows as samples and the columns as parameters "
                            "in whatever order the lab used. Header rows + "
                            "unit rows + banner rows above the data table "
                            "are common — skip them. Decimal values may use "
                            "comma or period.\n\n"
                            f"=== SPREADSHEET CONTENT ===\n{spreadsheet_text}"
                        ),
                    },
                    {"type": "text", "text": prompt},
                ]
            else:
                content_blocks = [
                    {
                        "type": "document" if content_type == "application/pdf" else "image",
                        "source": {
                            "type": source_type,
                            "media_type": media_type,
                            source_key: b64_data,
                        },
                    },
                    {"type": "text", "text": prompt},
                ]
            message = client.messages.create(
                model=_LAB_EXTRACTOR_MODEL,
                # 12k accommodates ~100-sample reports with 28 SOIL_PARAMS
                # each (each value pair ≈ 30 chars JSON). Hard ceiling
                # without adaptive thinking is 12000 × $15/M = $0.18,
                # still well within budget.
                max_tokens=12000,
                messages=[{"role": "user", "content": content_blocks}],
                # Lab extraction is structured value copying, not
                # reasoning — adaptive thinking just eats output tokens
                # before the JSON is emitted. effort=high gives careful
                # extraction without the thinking-budget tax.
                output_config={"effort": "high"},
            )
            break
        except _RETRYABLE_ERRORS as exc:
            last_err = exc
            if attempt == 0:
                logger.warning("Anthropic transient error on extract, retrying once: %s", exc)
                time.sleep(1.5)
                continue
            raise
    if message is None:
        raise last_err or RuntimeError("Extraction failed without a response")

    # Track usage
    input_tokens = message.usage.input_tokens if message.usage else 0
    output_tokens = message.usage.output_tokens if message.usage else 0
    # Sonnet pricing: $3/M input, $15/M output
    cost_usd = round((input_tokens * 3 + output_tokens * 15) / 1_000_000, 6)

    # Parse the response
    # Adaptive thinking puts ThinkingBlock(s) before the TextBlock(s)
    # in message.content. We need only the assistant's actual text
    # reply — iterate and concatenate `text`-typed blocks.
    response_text = "".join(
        getattr(block, "text", "") or ""
        for block in (message.content or [])
        if getattr(block, "type", None) == "text"
    ).strip()
    if not response_text:
        # No text block at all (truncated mid-thinking, etc.). Log the
        # cost we already paid and surface a clear error rather than
        # crashing on an empty parse.
        try:
            if user_id:
                sb = get_supabase_admin()
                sb.table("ai_usage").insert({
                    "user_id": user_id,
                    "operation": "extract_lab",
                    "model": _LAB_EXTRACTOR_MODEL,
                    "input_tokens": input_tokens,
                    "output_tokens": output_tokens,
                    "cost_usd": cost_usd,
                    "metadata": {
                        "status": "no_text_block",
                        "stop_reason": getattr(message, "stop_reason", None),
                        "block_types": [
                            getattr(b, "type", None) for b in (message.content or [])
                        ],
                    },
                }).execute()
        except Exception:
            logger.exception("ai_usage no-text-block log failed")
        raise ValueError(
            f"AI returned no text content "
            f"(stop_reason={getattr(message, 'stop_reason', None)!r}). "
            f"Likely cause: max_tokens reached during the thinking phase. "
            f"Try a smaller / cleaner file, or split the report."
        )
    # Strip markdown code fences if present (```json ... ``` or ``` ... ```)
    if response_text.startswith("```"):
        # Remove opening fence line (```json or ```)
        first_newline = response_text.find("\n")
        if first_newline > 0:
            response_text = response_text[first_newline + 1:]
        else:
            response_text = response_text[3:]
        # Remove closing fence
        if response_text.rstrip().endswith("```"):
            response_text = response_text.rstrip()[:-3].rstrip()

    try:
        extracted = json.loads(response_text)
    except json.JSONDecodeError:
        # Log the usage we already incurred — without this, a parse
        # failure on a $2 call leaves no trail in ai_usage and the cost
        # is invisible. The route's own ai_usage logger only fires on
        # success; this is the failure-path bookkeeper.
        try:
            if user_id:
                sb = get_supabase_admin()
                sb.table("ai_usage").insert({
                    "user_id": user_id,
                    "operation": "extract_lab",
                    "model": _LAB_EXTRACTOR_MODEL,
                    "input_tokens": input_tokens,
                    "output_tokens": output_tokens,
                    "cost_usd": cost_usd,
                    "metadata": {
                        "status": "json_parse_failed",
                        "response_excerpt": response_text[:500],
                        "stop_reason": getattr(message, "stop_reason", None),
                    },
                }).execute()
        except Exception:
            logger.exception("ai_usage failure-path log failed")
        raise ValueError(
            f"Failed to parse extraction result from AI "
            f"(stop_reason={getattr(message, 'stop_reason', None)!r}, "
            f"in={input_tokens} out={output_tokens} tokens, "
            f"~${cost_usd}): {response_text[:200]}"
        )

    # If we got a lab name, try to find/create template — keep
    # soil/leaf templates separate via department.
    detected_lab = extracted.get("lab_name", "")
    if detected_lab and not lab_template:
        lab_template = _get_lab_template(detected_lab, department=mode)

    # Handle both multi-sample and single-sample responses
    samples = extracted.get("samples", [])
    if not samples and extracted.get("values"):
        # Old single-sample format — wrap it
        samples = [{
            "sample_id": extracted.get("sample_id"),
            "block_name": None,
            "crop": None,
            "cultivar": None,
            "values": extracted["values"],
        }]

    # Save file + extraction result for future OCR training (fire-and-forget)
    try:
        _save_template_sample(file_bytes, content_type, detected_lab, extracted)
    except Exception:
        pass  # Never block the response for template saving

    return {
        "lab_name": detected_lab,
        "report_number": extracted.get("report_number"),
        "analysis_date": extracted.get("analysis_date"),
        "client": extracted.get("client"),
        "department": extracted.get("department"),
        "num_samples": len(samples),
        "confidence": extracted.get("confidence"),
        "ai_usage": {
            "input_tokens": input_tokens,
            "output_tokens": output_tokens,
            "cost_usd": cost_usd,
            "model": _LAB_EXTRACTOR_MODEL,
        },
        "samples": samples,
        # Keep backward-compatible single "values" for single-sample reports
        "values": samples[0]["values"] if len(samples) == 1 else {},
        "has_template": lab_template is not None,
        "template_sample_count": lab_template.get("sample_count", 0) if lab_template else 0,
    }


def learn_from_corrections(
    lab_name: str,
    original_values: dict[str, Any],
    corrected_values: dict[str, Any],
) -> dict[str, Any]:
    """Learn field mappings from user corrections.

    Compares original extracted values with user-corrected values to build
    a mapping of lab-specific field names to our standard parameter names.
    """
    if not lab_name:
        return {"updated": False, "reason": "No lab name provided"}

    sb = get_supabase_admin()

    # Get or create lab template
    result = sb.table("lab_templates").select("*").eq("lab_name", lab_name).execute()
    existing = result.data[0] if result.data else None

    field_mappings = existing["field_mappings"] if existing else {}

    # Find corrections: values that were changed or added
    changes = 0
    for param, corrected_val in corrected_values.items():
        original_val = original_values.get(param)
        if original_val != corrected_val and corrected_val is not None:
            # The user corrected this value — record the mapping
            # We store the original lab field name → our standard name
            field_mappings[param] = param
            changes += 1

    if existing:
        sb.table("lab_templates").update({
            "field_mappings": field_mappings,
            "sample_count": (existing.get("sample_count") or 0) + 1,
            "last_used": "now()",
        }).eq("id", existing["id"]).execute()
    else:
        sb.table("lab_templates").insert({
            "lab_name": lab_name,
            "field_mappings": field_mappings,
            "sample_count": 1,
        }).execute()

    return {
        "updated": True,
        "lab_name": lab_name,
        "corrections_recorded": changes,
        "total_mappings": len(field_mappings),
    }


def _get_lab_template(lab_name: str, department: str | None = None) -> dict | None:
    """Fetch a lab template by name (case-insensitive partial match).

    When `department` is given, prefer a template whose department matches
    so soil and leaf templates from the same lab don't collide. Falls back
    to any matching template if no exact dept match exists.
    """
    sb = get_supabase_admin()
    try:
        if department:
            result = (
                sb.table("lab_templates")
                .select("*")
                .ilike("lab_name", f"%{lab_name}%")
                .ilike("department", f"%{department}%")
                .limit(1)
                .execute()
            )
            if result.data:
                return result.data[0]
        result = sb.table("lab_templates").select("*").ilike("lab_name", f"%{lab_name}%").limit(1).execute()
        return result.data[0] if result.data else None
    except Exception:
        return None


def _save_template_sample(
    file_bytes: bytes,
    content_type: str,
    lab_name: str,
    extraction_result: dict,
) -> None:
    """Save the original file and extraction result to Supabase Storage + lab_templates.

    Stored for future OCR engine training. Each file gets a unique path under
    lab-reports/{lab_name_slug}/{timestamp}.{ext}
    """
    import hashlib
    from datetime import datetime

    if not lab_name:
        return

    sb = get_supabase_admin()

    # Generate file path
    dept = (extraction_result.get("department") or "unknown").lower().replace(" ", "-")[:20]
    slug = lab_name.lower().replace(" ", "-").replace("/", "-")[:50]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ext = {
        "application/pdf": "pdf",
        "image/jpeg": "jpg",
        "image/png": "png",
        "image/webp": "webp",
    }.get(content_type, "bin")
    file_hash = hashlib.md5(file_bytes).hexdigest()[:8]
    storage_path = f"{slug}/{dept}/{ts}_{file_hash}.{ext}"

    # Upload file to storage
    sb.storage.from_("lab-reports").upload(
        path=storage_path,
        file=file_bytes,
        file_options={"content-type": content_type},
    )

    # Update lab_templates with the file reference and extraction layout
    # Use lab_name + department as key so soil and leaf get separate templates
    template_name = f"{lab_name} ({dept})" if dept != "unknown" else lab_name
    result = sb.table("lab_templates").select("*").eq("lab_name", template_name).limit(1).execute()
    if not result.data:
        # Fall back to matching without department
        result = sb.table("lab_templates").select("*").ilike("lab_name", f"%{lab_name}%").limit(1).execute()

    # Extract layout hints from the result (column order, header names, etc.)
    layout_data = {
        "department": extraction_result.get("department"),
        "num_samples": len(extraction_result.get("samples", [])),
        "column_headers": [],
        "sample_fields": [],
    }
    samples = extraction_result.get("samples", [])
    if samples:
        # Record what fields were found (this is the "layout" of the report)
        layout_data["sample_fields"] = list(samples[0].get("values", {}).keys())
        layout_data["has_block_name"] = any(s.get("block_name") for s in samples)
        layout_data["has_crop"] = any(s.get("crop") for s in samples)
        layout_data["has_cultivar"] = any(s.get("cultivar") for s in samples)

    if result.data:
        existing = result.data[0]
        sample_files = existing.get("sample_files") or []
        sample_files.append(storage_path)
        # Keep max 20 samples per lab (oldest get pushed out)
        if len(sample_files) > 20:
            sample_files = sample_files[-20:]

        sb.table("lab_templates").update({
            "sample_files": sample_files,
            "layout_data": layout_data,
            "sample_count": (existing.get("sample_count") or 0) + 1,
            "last_used": "now()",
        }).eq("id", existing["id"]).execute()
    else:
        sb.table("lab_templates").insert({
            "lab_name": template_name,
            "field_mappings": {},
            "sample_files": [storage_path],
            "layout_data": layout_data,
            "sample_count": 1,
        }).execute()
